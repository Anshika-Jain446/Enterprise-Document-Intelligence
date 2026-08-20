import os
import io
import zipfile
import hashlib
import mimetypes
import xml.etree.ElementTree as ET

import fitz  # PyMuPDF
import pdfplumber
import pandas as pd
from openpyxl import load_workbook

# Optional OCR
try:
    import pytesseract
    from PIL import Image

    OCR_AVAILABLE = True
except Exception:
    OCR_AVAILABLE = False

# Optional Gemini Vision
try:
    from google import genai
    from google.genai import types

    GEMINI_AVAILABLE = True
except Exception:
    GEMINI_AVAILABLE = False


class DocumentExtractor:

    def __init__(self, file_path: str):

        if not os.path.isfile(file_path):
            raise FileNotFoundError(
                f"Document file not found: {file_path}"
            )

        self.file_path = file_path
        self.pdf_path = file_path

        self.extension = os.path.splitext(
            file_path
        )[1].lower()

        self.file_name = os.path.basename(
            file_path
        )

        self.base_name = os.path.splitext(
            self.file_name
        )[0]

        self.image_output_dir = os.path.join(
            "outputs",
            "images",
            self.base_name
        )

        self.visual_output_dir = os.path.join(
            "outputs",
            "visuals",
            self.base_name
        )

        os.makedirs(
            self.image_output_dir,
            exist_ok=True
        )

        os.makedirs(
            self.visual_output_dir,
            exist_ok=True
        )

        # ======================================================
        # IMAGE UNDERSTANDING CONFIGURATION
        # ======================================================

        self.enable_ocr = (
            os.getenv(
                "ENABLE_OCR",
                "true"
            ).lower()
            == "true"
        )

        self.enable_vision = (
            os.getenv(
                "ENABLE_VISION",
                "true"
            ).lower()
            == "true"
        )

        self.vision_model = os.getenv(
            "VISION_MODEL",
            "gemini-2.5-flash"
        )

        self.vision_client = None

        if (
            self.enable_vision
            and GEMINI_AVAILABLE
        ):

            api_key = os.getenv(
                "GEMINI_API_KEY"
            )

            if api_key:

                try:

                    self.vision_client = (
                        genai.Client(
                            api_key=api_key
                        )
                    )

                except Exception:

                    self.vision_client = None

    # ==========================================================
    # HELPERS
    # ==========================================================

    def _safe_text(self, value):

        if value is None:
            return ""

        return str(value).strip()

    def _clean_ocr_text(self, text):

        if not text:
            return ""

        lines = []

        for line in text.splitlines():

            line = line.strip()

            if line:
                lines.append(line)

        return "\n".join(lines)

    def _image_to_pil(self, image_path):

        if not OCR_AVAILABLE:
            return None

        try:

            return Image.open(
                image_path
            ).convert("RGB")

        except Exception:

            return None

    # ==========================================================
    # OCR
    # ==========================================================

    def _run_ocr(self, image_path):

        if not self.enable_ocr:
            return ""

        if not OCR_AVAILABLE:
            return ""

        image = self._image_to_pil(
            image_path
        )

        if image is None:
            return ""

        try:

            text = pytesseract.image_to_string(
                image
            )

            return self._clean_ocr_text(
                text
            )

        except Exception:

            return ""

    # ==========================================================
    # GEMINI VISION
    # ==========================================================

    def _run_vision_analysis(
        self,
        image_path,
        page_number=None,
        image_number=None,
    ):

        if not self.enable_vision:
            return ""

        if self.vision_client is None:
            return ""

        try:

            with open(
                image_path,
                "rb"
            ) as image_file:

                image_bytes = (
                    image_file.read()
                )

            prompt = f"""
You are the visual-document analysis component
of an Enterprise Agentic RAG system.

Analyze this image carefully.

The image may contain:
- charts
- graphs
- tables
- diagrams
- flowcharts
- architecture diagrams
- screenshots
- forms
- scanned documents
- photographs
- technical drawings
- handwritten content
- business visuals
- equations
- labels
- captions

Extract and explain information that could be
important for answering future questions about
the source document.

Do NOT hallucinate information.

Separate what is clearly visible from uncertain
interpretations.

Return a structured textual analysis containing:

1. IMAGE TYPE
2. VISIBLE TEXT
3. MAIN SUBJECT
4. DETAILED DESCRIPTION
5. IMPORTANT ENTITIES / LABELS
6. NUMBERS / VALUES
7. TABLE OR CHART INFORMATION
8. RELATIONSHIPS / CONNECTIONS
9. TECHNICAL OR BUSINESS MEANING
10. IMPORTANT FACTS FOR QUESTION ANSWERING

If the image is a chart, explain:
- title
- axes
- legend
- categories
- values
- trends
- comparisons

If it is a diagram or flowchart, explain:
- components
- labels
- arrows
- relationships
- sequence/process

If it is a screenshot, explain:
- visible application/interface
- important fields
- messages
- values
- controls

If it is a scanned page, extract as much readable
content as possible.

Do not invent missing information.

Source document:
{self.file_name}

Page:
{page_number if page_number is not None else "Unknown"}

Image:
{image_number if image_number is not None else "Unknown"}
"""

            response = (
                self.vision_client.models.generate_content(
                    model=self.vision_model,
                    contents=[
                        types.Part.from_bytes(
                            data=image_bytes,
                            mime_type=(
                                mimetypes.guess_type(
                                    image_path
                                )[0]
                                or "image/png"
                            ),
                        ),
                        prompt,
                    ],
                )
            )

            if response is None:
                return ""

            text = getattr(
                response,
                "text",
                ""
            )

            return self._safe_text(
                text
            )

        except Exception:

            return ""

    # ==========================================================
    # COMBINED IMAGE UNDERSTANDING
    # ==========================================================

    def _analyze_image(
        self,
        image_path,
        page_number=None,
        image_number=None,
    ):

        ocr_text = self._run_ocr(
            image_path
        )

        vision_description = (
            self._run_vision_analysis(
                image_path=image_path,
                page_number=page_number,
                image_number=image_number,
            )
        )

        return {
            "ocr_text": ocr_text,
            "image_description": (
                vision_description
            ),
        }

    # ==========================================================
    # METADATA
    # ==========================================================

    def extract_metadata(self):

        try:

            metadata = {
                "file_name": self.file_name,
                "file_type": self.extension,
                "title": self.base_name,
                "author": "Unknown",
                "subject": "Unknown",
                "creator": "Unknown",
                "producer": "Unknown",
                "pages": 1,
            }

            # --------------------------------------------------
            # PDF
            # --------------------------------------------------

            if self.extension == ".pdf":

                with fitz.open(
                    self.file_path
                ) as doc:

                    pdf_metadata = (
                        doc.metadata or {}
                    )

                    metadata.update({
                        "title":
                            pdf_metadata.get(
                                "title"
                            )
                            or self.base_name,

                        "author":
                            pdf_metadata.get(
                                "author"
                            )
                            or "Unknown",

                        "subject":
                            pdf_metadata.get(
                                "subject"
                            )
                            or "Unknown",

                        "creator":
                            pdf_metadata.get(
                                "creator"
                            )
                            or "Unknown",

                        "producer":
                            pdf_metadata.get(
                                "producer"
                            )
                            or "Unknown",

                        "pages":
                            len(doc),

                        "document_type":
                            "PDF Document",
                    })

            # --------------------------------------------------
            # DOCX
            # --------------------------------------------------

            elif self.extension == ".docx":

                metadata["document_type"] = (
                    "Microsoft Word Document"
                )

                with zipfile.ZipFile(
                    self.file_path,
                    "r"
                ) as archive:

                    try:

                        xml_data = archive.read(
                            "docProps/core.xml"
                        )

                        root = ET.fromstring(
                            xml_data
                        )

                        namespaces = {
                            "dc":
                                "http://purl.org/dc/elements/1.1/"
                        }

                        title = root.find(
                            "dc:title",
                            namespaces
                        )

                        creator = root.find(
                            "dc:creator",
                            namespaces
                        )

                        subject = root.find(
                            "dc:subject",
                            namespaces
                        )

                        if (
                            title is not None
                            and title.text
                        ):
                            metadata["title"] = (
                                title.text
                            )

                        if (
                            creator is not None
                            and creator.text
                        ):
                            metadata["author"] = (
                                creator.text
                            )

                        if (
                            subject is not None
                            and subject.text
                        ):
                            metadata["subject"] = (
                                subject.text
                            )

                    except KeyError:
                        pass

            # --------------------------------------------------
            # PPTX
            # --------------------------------------------------

            elif self.extension == ".pptx":

                metadata["document_type"] = (
                    "PowerPoint Presentation"
                )

                with zipfile.ZipFile(
                    self.file_path,
                    "r"
                ) as archive:

                    slide_files = [
                        name
                        for name in archive.namelist()
                        if name.startswith(
                            "ppt/slides/slide"
                        )
                        and name.endswith(
                            ".xml"
                        )
                    ]

                    metadata["pages"] = len(
                        slide_files
                    )

            # --------------------------------------------------
            # XLSX
            # --------------------------------------------------

            elif self.extension == ".xlsx":

                metadata["document_type"] = (
                    "Excel Workbook"
                )

                workbook = load_workbook(
                    self.file_path,
                    read_only=True,
                    data_only=True,
                )

                metadata["sheets"] = len(
                    workbook.sheetnames
                )

                metadata["sheet_names"] = (
                    workbook.sheetnames
                )

                metadata["pages"] = max(
                    1,
                    len(workbook.sheetnames)
                )

                workbook.close()

            # --------------------------------------------------
            # CSV
            # --------------------------------------------------

            elif self.extension == ".csv":

                metadata["document_type"] = (
                    "CSV Dataset"
                )

            # --------------------------------------------------
            # TXT / MD
            # --------------------------------------------------

            elif self.extension in [
                ".txt",
                ".md",
            ]:

                metadata["document_type"] = (
                    "Markdown Document"
                    if self.extension == ".md"
                    else "Text Document"
                )

            return metadata

        except Exception as e:

            return {
                "file_name": self.file_name,
                "file_type": self.extension,
                "pages": 1,
                "error":
                    f"Metadata extraction failed: {e}",
            }

    # ==========================================================
    # PDF TEXT
    # ==========================================================

    def _extract_pdf_text(self):

        pages = []
        text_chunks = []

        with fitz.open(
            self.file_path
        ) as doc:

            for page_num, page in enumerate(
                doc,
                start=1,
            ):

                text = (
                    page.get_text(
                        "text"
                    )
                    or ""
                )

                pages.append({
                    "page": page_num,
                    "text": text,
                })

                text_chunks.append(
                    text
                )

        return (
            "\n".join(text_chunks),
            pages,
        )

    # ==========================================================
    # DOCX TEXT
    # ==========================================================

    def _extract_docx_text(self):

        text_parts = []

        with zipfile.ZipFile(
            self.file_path,
            "r"
        ) as archive:

            xml_data = archive.read(
                "word/document.xml"
            )

            root = ET.fromstring(
                xml_data
            )

            namespace = {
                "w":
                    "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
            }

            paragraphs = root.findall(
                ".//w:p",
                namespace
            )

            for paragraph in paragraphs:

                words = []

                for text_node in paragraph.findall(
                    ".//w:t",
                    namespace
                ):

                    if text_node.text:
                        words.append(
                            text_node.text
                        )

                paragraph_text = "".join(
                    words
                ).strip()

                if paragraph_text:
                    text_parts.append(
                        paragraph_text
                    )

        full_text = "\n".join(
            text_parts
        )

        pages = [{
            "page": 1,
            "text": full_text,
        }]

        return (
            full_text,
            pages,
        )

    # ==========================================================
    # TXT / MARKDOWN
    # ==========================================================

    def _extract_text_file(self):

        with open(
            self.file_path,
            "r",
            encoding="utf-8",
            errors="ignore",
        ) as file:

            text = file.read()

        pages = [{
            "page": 1,
            "text": text,
        }]

        return (
            text,
            pages,
        )

    # ==========================================================
    # CSV
    # ==========================================================

    def _extract_csv(self):

        tables = []

        dataframe = pd.read_csv(
            self.file_path
        )

        table_records = (
            dataframe
            .fillna("")
            .to_dict(
                orient="records"
            )
        )

        tables.append({
            "page": 1,
            "rows": len(dataframe),
            "columns": len(dataframe.columns),
            "table": table_records,
            "source": "CSV",
        })

        text = dataframe.to_csv(
            index=False
        )

        pages = [{
            "page": 1,
            "text": text,
        }]

        return (
            text,
            pages,
            tables,
        )

    # ==========================================================
    # XLSX
    # ==========================================================

    def _extract_xlsx(self):

        tables = []
        text_parts = []

        workbook = load_workbook(
            self.file_path,
            data_only=True,
        )

        for sheet in workbook.worksheets:

            rows = list(
                sheet.iter_rows(
                    values_only=True
                )
            )

            if not rows:
                continue

            headers = []

            for index, value in enumerate(
                rows[0]
            ):

                if value is None:
                    headers.append(
                        f"Column_{index}"
                    )

                else:
                    headers.append(
                        str(value)
                    )

            records = []

            for row in rows[1:]:

                record = {}

                for index, value in enumerate(
                    row
                ):

                    if index < len(headers):

                        record[
                            headers[index]
                        ] = (
                            ""
                            if value is None
                            else value
                        )

                records.append(
                    record
                )

            tables.append({
                "page": sheet.title,
                "sheet": sheet.title,
                "rows": len(records),
                "columns": len(headers),
                "table": records,
                "source": "XLSX",
            })

            text_parts.append(
                f"Sheet: {sheet.title}"
            )

            for row in rows:

                text_parts.append(
                    " | ".join(
                        ""
                        if value is None
                        else str(value)
                        for value in row
                    )
                )

        workbook.close()

        text = "\n".join(
            text_parts
        )

        pages = [{
            "page": 1,
            "text": text,
        }]

        return (
            text,
            pages,
            tables,
        )

    # ==========================================================
    # PPTX TEXT
    # ==========================================================

    def _extract_pptx_text(self):

        pages = []
        text_parts = []

        namespace = {
            "a":
                "http://schemas.openxmlformats.org/drawingml/2006/main"
        }

        with zipfile.ZipFile(
            self.file_path,
            "r"
        ) as archive:

            slide_files = [
                name
                for name in archive.namelist()
                if name.startswith(
                    "ppt/slides/slide"
                )
                and name.endswith(
                    ".xml"
                )
            ]

            def slide_number(name):

                filename = os.path.basename(
                    name
                )

                number = (
                    os.path.splitext(
                        filename
                    )[0]
                    .replace(
                        "slide",
                        ""
                    )
                )

                try:
                    return int(number)

                except ValueError:
                    return 0

            slide_files.sort(
                key=slide_number
            )

            for slide_file in slide_files:

                slide_num = slide_number(
                    slide_file
                )

                xml_data = archive.read(
                    slide_file
                )

                root = ET.fromstring(
                    xml_data
                )

                texts = []

                for node in root.findall(
                    ".//a:t",
                    namespace
                ):

                    if node.text:
                        texts.append(
                            node.text
                        )

                slide_text = "\n".join(
                    texts
                )

                pages.append({
                    "page": slide_num,
                    "text": slide_text,
                })

                text_parts.append(
                    f"Slide {slide_num}"
                )

                text_parts.append(
                    slide_text
                )

        return (
            "\n".join(text_parts),
            pages,
        )

    # ==========================================================
    # MAIN TEXT EXTRACTION
    # ==========================================================

    def extract_text(self):

        try:

            if self.extension == ".pdf":

                return (
                    self._extract_pdf_text()
                )

            elif self.extension == ".docx":

                return (
                    self._extract_docx_text()
                )

            elif self.extension in [
                ".txt",
                ".md",
            ]:

                return (
                    self._extract_text_file()
                )

            elif self.extension == ".csv":

                text, pages, _ = (
                    self._extract_csv()
                )

                return (
                    text,
                    pages,
                )

            elif self.extension == ".xlsx":

                text, pages, _ = (
                    self._extract_xlsx()
                )

                return (
                    text,
                    pages,
                )

            elif self.extension == ".pptx":

                return (
                    self._extract_pptx_text()
                )

            else:

                return (
                    "",
                    [{
                        "error":
                            "Unsupported file type: "
                            f"{self.extension}"
                    }],
                )

        except Exception as e:

            return (
                "",
                [{
                    "error":
                        f"Text extraction failed: {e}"
                }],
            )

    # ==========================================================
    # PDF TABLES
    # ==========================================================

    def _extract_pdf_tables(self):

        tables = []

        with pdfplumber.open(
            self.file_path
        ) as pdf:

            for page_num, page in enumerate(
                pdf.pages,
                start=1,
            ):

                extracted_tables = (
                    page.extract_tables()
                    or []
                )

                for table in extracted_tables:

                    if len(table) < 2:
                        continue

                    headers = []

                    for index, column in enumerate(
                        table[0]
                    ):

                        header = (
                            str(column).strip()
                            if column
                            else f"Column_{index}"
                        )

                        headers.append(
                            header
                        )

                    try:

                        dataframe = pd.DataFrame(
                            table[1:],
                            columns=headers,
                        )

                        tables.append({
                            "page": page_num,
                            "rows": len(dataframe),
                            "columns":
                                len(
                                    dataframe.columns
                                ),
                            "table":
                                dataframe
                                .fillna("")
                                .to_dict(
                                    orient="records"
                                ),
                            "source": "PDF",
                        })

                    except Exception as error:

                        tables.append({
                            "page": page_num,
                            "error": str(error),
                            "source": "PDF",
                        })

        return tables

    # ==========================================================
    # TABLES
    # ==========================================================

    def extract_tables(self):

        try:

            if self.extension == ".pdf":

                return (
                    self._extract_pdf_tables()
                )

            elif self.extension == ".csv":

                _, _, tables = (
                    self._extract_csv()
                )

                return tables

            elif self.extension == ".xlsx":

                _, _, tables = (
                    self._extract_xlsx()
                )

                return tables

            else:

                return []

        except Exception as e:

            return [{
                "error":
                    f"Table extraction failed: {e}"
            }]

    # ==========================================================
    # PDF EMBEDDED IMAGES
    # ==========================================================

    def _extract_pdf_images(self):

        images = []

        with fitz.open(
            self.file_path
        ) as doc:

            for page_num, page in enumerate(
                doc,
                start=1,
            ):

                page_images = page.get_images(
                    full=True
                )

                image_info = []

                for index, img in enumerate(
                    page_images,
                    start=1,
                ):

                    xref = img[0]

                    try:

                        image_data = (
                            doc.extract_image(
                                xref
                            )
                        )

                        image_bytes = (
                            image_data["image"]
                        )

                        image_ext = (
                            image_data["ext"]
                        )

                        image_filename = (
                            f"page_{page_num}_"
                            f"image_{index}."
                            f"{image_ext}"
                        )

                        image_path = os.path.join(
                            self.image_output_dir,
                            image_filename,
                        )

                        with open(
                            image_path,
                            "wb"
                        ) as image_file:

                            image_file.write(
                                image_bytes
                            )

                        analysis = (
                            self._analyze_image(
                                image_path=image_path,
                                page_number=page_num,
                                image_number=index,
                            )
                        )

                        image_info.append({
                            "image_no": index,
                            "xref": xref,
                            "width": img[2],
                            "height": img[3],
                            "bits_per_component":
                                img[4],
                            "colorspace": img[5],
                            "format": image_ext,
                            "path": image_path,
                            "ocr_text":
                                analysis.get(
                                    "ocr_text",
                                    ""
                                ),
                            "image_description":
                                analysis.get(
                                    "image_description",
                                    ""
                                ),
                        })

                    except Exception as image_error:

                        image_info.append({
                            "image_no": index,
                            "xref": xref,
                            "error":
                                str(image_error),
                        })

                images.append({
                    "page": page_num,
                    "count": len(page_images),
                    "images": image_info,
                })

        return images

    # ==========================================================
    # DOCX IMAGES
    # ==========================================================

    def _extract_docx_images(self):

        images = []

        with zipfile.ZipFile(
            self.file_path,
            "r"
        ) as archive:

            media_files = [
                name
                for name in archive.namelist()
                if name.startswith(
                    "word/media/"
                )
            ]

            for index, media_file in enumerate(
                media_files,
                start=1,
            ):

                extension = os.path.splitext(
                    media_file
                )[1].lower()

                output_name = (
                    f"docx_image_{index}"
                    f"{extension}"
                )

                output_path = os.path.join(
                    self.image_output_dir,
                    output_name,
                )

                with open(
                    output_path,
                    "wb"
                ) as output:

                    output.write(
                        archive.read(
                            media_file
                        )
                    )

                analysis = (
                    self._analyze_image(
                        image_path=output_path,
                        page_number=1,
                        image_number=index,
                    )
                )

                images.append({
                    "image_no": index,
                    "source": media_file,
                    "path": output_path,
                    "ocr_text":
                        analysis.get(
                            "ocr_text",
                            ""
                        ),
                    "image_description":
                        analysis.get(
                            "image_description",
                            ""
                        ),
                })

        return [{
            "page": 1,
            "count": len(images),
            "images": images,
        }]

    # ==========================================================
    # PPTX IMAGES
    # ==========================================================

    def _extract_pptx_images(self):

        images = []

        with zipfile.ZipFile(
            self.file_path,
            "r"
        ) as archive:

            media_files = [
                name
                for name in archive.namelist()
                if name.startswith(
                    "ppt/media/"
                )
            ]

            for index, media_file in enumerate(
                media_files,
                start=1,
            ):

                extension = os.path.splitext(
                    media_file
                )[1].lower()

                output_name = (
                    f"pptx_image_{index}"
                    f"{extension}"
                )

                output_path = os.path.join(
                    self.image_output_dir,
                    output_name,
                )

                with open(
                    output_path,
                    "wb"
                ) as output:

                    output.write(
                        archive.read(
                            media_file
                        )
                    )

                analysis = (
                    self._analyze_image(
                        image_path=output_path,
                        page_number=1,
                        image_number=index,
                    )
                )

                images.append({
                    "image_no": index,
                    "source": media_file,
                    "path": output_path,
                    "ocr_text":
                        analysis.get(
                            "ocr_text",
                            ""
                        ),
                    "image_description":
                        analysis.get(
                            "image_description",
                            ""
                        ),
                })

        return [{
            "page": 1,
            "count": len(images),
            "images": images,
        }]

    # ==========================================================
    # MAIN IMAGE EXTRACTION
    # ==========================================================

    def extract_images(self):

        try:

            if self.extension == ".pdf":

                return (
                    self._extract_pdf_images()
                )

            elif self.extension == ".docx":

                return (
                    self._extract_docx_images()
                )

            elif self.extension == ".pptx":

                return (
                    self._extract_pptx_images()
                )

            else:

                return []

        except Exception as e:

            return [{
                "error":
                    f"Image extraction failed: {e}"
            }]

    # ==========================================================
    # IMAGE KNOWLEDGE TEXT
    #
    # THIS IS IMPORTANT FOR AGENTIC RAG.
    # It converts image understanding into text
    # that can be chunked + embedded.
    # ==========================================================

    def build_image_knowledge_text(
        self,
        images=None,
    ):

        if images is None:
            images = self.extract_images()

        knowledge_parts = []

        for page_data in images:

            if not isinstance(
                page_data,
                dict
            ):
                continue

            page_number = page_data.get(
                "page",
                1
            )

            for image in page_data.get(
                "images",
                []
            ):

                if not isinstance(
                    image,
                    dict
                ):
                    continue

                image_number = image.get(
                    "image_no",
                    "?"
                )

                ocr_text = self._safe_text(
                    image.get(
                        "ocr_text",
                        ""
                    )
                )

                description = self._safe_text(
                    image.get(
                        "image_description",
                        ""
                    )
                )

                if not ocr_text and not description:
                    continue

                knowledge_parts.append(
                    "\n".join([
                        "IMAGE KNOWLEDGE",
                        f"Source: {self.file_name}",
                        f"Page: {page_number}",
                        f"Image: {image_number}",
                        "",
                        "OCR TEXT:",
                        ocr_text
                        or "No readable OCR text detected.",
                        "",
                        "VISUAL ANALYSIS:",
                        description
                        or "No visual analysis available.",
                    ])
                )

        return "\n\n".join(
            knowledge_parts
        )

    # ==========================================================
    # IMAGE PAGES FOR PDF
    #
    # This captures full-page visuals, diagrams,
    # flowcharts and vector drawings.
    # ==========================================================

    def _extract_pdf_visuals(self):

        visuals = []

        with fitz.open(
            self.file_path
        ) as doc:

            for page_num, page in enumerate(
                doc,
                start=1,
            ):

                embedded_image_count = len(
                    page.get_images(
                        full=True
                    )
                )

                try:

                    drawing_count = len(
                        page.get_drawings()
                    )

                except Exception:

                    drawing_count = 0

                has_visual_content = (
                    embedded_image_count > 0
                    or drawing_count > 0
                )

                if not has_visual_content:
                    continue

                output_name = (
                    f"page_{page_num}_visual.png"
                )

                output_path = os.path.join(
                    self.visual_output_dir,
                    output_name,
                )

                try:

                    matrix = fitz.Matrix(
                        2.0,
                        2.0
                    )

                    pixmap = page.get_pixmap(
                        matrix=matrix,
                        alpha=False,
                    )

                    pixmap.save(
                        output_path
                    )

                    # Full-page visual understanding
                    visual_analysis = (
                        self._analyze_image(
                            image_path=output_path,
                            page_number=page_num,
                            image_number="full_page",
                        )
                    )

                    visuals.append({
                        "page": page_num,
                        "type":
                            "PDF rendered visual",
                        "path":
                            output_path,
                        "embedded_images":
                            embedded_image_count,
                        "drawings":
                            drawing_count,
                        "ocr_text":
                            visual_analysis.get(
                                "ocr_text",
                                ""
                            ),
                        "image_description":
                            visual_analysis.get(
                                "image_description",
                                ""
                            ),
                    })

                except Exception as render_error:

                    visuals.append({
                        "page": page_num,
                        "type":
                            "PDF rendered visual",
                        "error":
                            str(render_error),
                        "embedded_images":
                            embedded_image_count,
                        "drawings":
                            drawing_count,
                    })

        return visuals

    # ==========================================================
    # VISUAL EXTRACTION
    # ==========================================================

    def extract_visuals(self):

        try:

            if self.extension == ".pdf":

                return (
                    self._extract_pdf_visuals()
                )

            return []

        except Exception as e:

            return [{
                "error":
                    f"Visual extraction failed: {e}"
            }]

    # ==========================================================
    # VISUAL KNOWLEDGE TEXT
    # ==========================================================

    def build_visual_knowledge_text(
        self,
        visuals=None,
    ):

        if visuals is None:
            visuals = self.extract_visuals()

        knowledge_parts = []

        for visual in visuals:

            if not isinstance(
                visual,
                dict
            ):
                continue

            if not visual.get(
                "path"
            ):
                continue

            page_number = visual.get(
                "page",
                "?"
            )

            ocr_text = self._safe_text(
                visual.get(
                    "ocr_text",
                    ""
                )
            )

            description = self._safe_text(
                visual.get(
                    "image_description",
                    ""
                )
            )

            if not ocr_text and not description:
                continue

            knowledge_parts.append(
                "\n".join([
                    "FULL PAGE VISUAL KNOWLEDGE",
                    f"Source: {self.file_name}",
                    f"Page: {page_number}",
                    "",
                    "OCR TEXT:",
                    ocr_text
                    or "No readable OCR text detected.",
                    "",
                    "VISUAL ANALYSIS:",
                    description
                    or "No visual analysis available.",
                ])
            )

        return "\n\n".join(
            knowledge_parts
        )

    # ==========================================================
    # COMPLETE MULTIMODAL KNOWLEDGE
    # ==========================================================

    def build_multimodal_knowledge(
        self,
        text,
        tables=None,
        images=None,
        visuals=None,
    ):

        if tables is None:
            tables = self.extract_tables()

        if images is None:
            images = self.extract_images()

        if visuals is None:
            visuals = self.extract_visuals()

        parts = []

        # ------------------------------------------------------
        # NORMAL DOCUMENT TEXT
        # ------------------------------------------------------

        if text:

            parts.append(
                f"""
DOCUMENT TEXT

Source: {self.file_name}

{text}
""".strip()
            )

        # ------------------------------------------------------
        # TABLE KNOWLEDGE
        # ------------------------------------------------------

        if tables:

            for index, table in enumerate(
                tables,
                start=1,
            ):

                if not isinstance(
                    table,
                    dict
                ):
                    continue

                if "table" not in table:
                    continue

                page = table.get(
                    "page",
                    "?"
                )

                table_text = []

                table_text.append(
                    "TABLE KNOWLEDGE"
                )

                table_text.append(
                    f"Source: {self.file_name}"
                )

                table_text.append(
                    f"Page/Sheet: {page}"
                )

                table_text.append("")

                dataframe = pd.DataFrame(
                    table["table"]
                )

                table_text.append(
                    dataframe.to_string(
                        index=False
                    )
                )

                parts.append(
                    "\n".join(table_text)
                )

        # ------------------------------------------------------
        # IMAGE KNOWLEDGE
        # ------------------------------------------------------

        image_knowledge = (
            self.build_image_knowledge_text(
                images
            )
        )

        if image_knowledge:

            parts.append(
                image_knowledge
            )

        # ------------------------------------------------------
        # FULL PAGE VISUAL KNOWLEDGE
        # ------------------------------------------------------

        visual_knowledge = (
            self.build_visual_knowledge_text(
                visuals
            )
        )

        if visual_knowledge:

            parts.append(
                visual_knowledge
            )

        return "\n\n".join(
            parts
        )

    # ==========================================================
    # STATISTICS
    # ==========================================================

    def document_statistics(
        self,
        text,
        tables=None,
        images=None,
        visuals=None,
    ):

        if tables is None:
            tables = self.extract_tables()

        if images is None:
            images = self.extract_images()

        if visuals is None:
            visuals = self.extract_visuals()

        total_images = 0

        analyzed_images = 0
        ocr_images = 0
        vision_images = 0

        for item in images:

            if not isinstance(
                item,
                dict
            ):
                continue

            total_images += item.get(
                "count",
                0
            )

            for image in item.get(
                "images",
                []
            ):

                if not isinstance(
                    image,
                    dict
                ):
                    continue

                ocr = self._safe_text(
                    image.get(
                        "ocr_text",
                        ""
                    )
                )

                vision = self._safe_text(
                    image.get(
                        "image_description",
                        ""
                    )
                )

                if ocr:
                    ocr_images += 1

                if vision:
                    vision_images += 1

                if ocr or vision:
                    analyzed_images += 1

        real_tables = [
            table
            for table in tables
            if (
                isinstance(
                    table,
                    dict
                )
                and "table" in table
            )
        ]

        visual_count = len([
            item
            for item in visuals
            if (
                isinstance(
                    item,
                    dict
                )
                and "path" in item
            )
        ])

        return {
            "words":
                len(
                    text.split()
                ),

            "characters":
                len(text),

            "lines":
                len(
                    text.splitlines()
                ),

            "tables":
                len(real_tables),

            "images":
                total_images,

            "analyzed_images":
                analyzed_images,

            "ocr_images":
                ocr_images,

            "vision_images":
                vision_images,

            "visuals":
                visual_count,

            "ocr_available":
                OCR_AVAILABLE,

            "vision_available":
                self.vision_client is not None,
        }

    # ==========================================================
    # COMPLETE DOCUMENT EXTRACTION
    # ==========================================================

    def extract_document(self):

        # ------------------------------------------------------
        # TEXT
        # ------------------------------------------------------

        text, pages = (
            self.extract_text()
        )

        # ------------------------------------------------------
        # TABLES
        # ------------------------------------------------------

        tables = (
            self.extract_tables()
        )

        # ------------------------------------------------------
        # EMBEDDED IMAGES
        # ------------------------------------------------------

        images = (
            self.extract_images()
        )

        # ------------------------------------------------------
        # PDF VISUAL PAGES
        # ------------------------------------------------------

        visuals = (
            self.extract_visuals()
        )

        # ------------------------------------------------------
        # METADATA
        # ------------------------------------------------------

        metadata = (
            self.extract_metadata()
        )

        # ------------------------------------------------------
        # MULTIMODAL KNOWLEDGE
        #
        # This is the important new field.
        # Your chunking/vector DB layer can index this.
        # ------------------------------------------------------

        multimodal_knowledge = (
            self.build_multimodal_knowledge(
                text=text,
                tables=tables,
                images=images,
                visuals=visuals,
            )
        )

        # ------------------------------------------------------
        # STATISTICS
        # ------------------------------------------------------

        statistics = (
            self.document_statistics(
                text=text,
                tables=tables,
                images=images,
                visuals=visuals,
            )
        )

        # ------------------------------------------------------
        # RETURN
        # ------------------------------------------------------

        return {
            "metadata": metadata,

            "statistics": statistics,

            "text": text,

            "pages": pages,

            "tables": tables,

            "images": images,

            "visuals": visuals,

            # New multimodal searchable knowledge.
            "multimodal_knowledge":
                multimodal_knowledge,
        }


# ==============================================================
# TEST
# ==============================================================

if __name__ == "__main__":

    file_path = "sample.pdf"

    try:

        extractor = DocumentExtractor(
            file_path
        )

        result = (
            extractor.extract_document()
        )

        print(
            "\n========== METADATA =========="
        )

        print(
            result["metadata"]
        )

        print(
            "\n========== STATISTICS =========="
        )

        print(
            result["statistics"]
        )

        print(
            "\n========== TABLES =========="
        )

        print(
            f"Tables Found: "
            f"{len(result['tables'])}"
        )

        print(
            "\n========== IMAGES =========="
        )

        total_images = sum(
            page.get(
                "count",
                0
            )
            for page in result["images"]
            if isinstance(
                page,
                dict
            )
        )

        print(
            f"Total Images: "
            f"{total_images}"
        )

        print(
            "\n========== IMAGE ANALYSIS =========="
        )

        for page in result["images"]:

            if not isinstance(
                page,
                dict
            ):
                continue

            for image in page.get(
                "images",
                []
            ):

                if not isinstance(
                    image,
                    dict
                ):
                    continue

                print(
                    f"\nPage: "
                    f"{page.get('page')}"
                )

                print(
                    f"Image: "
                    f"{image.get('image_no')}"
                )

                print(
                    "OCR:"
                )

                print(
                    image.get(
                        "ocr_text",
                        ""
                    )[:1000]
                )

                print(
                    "\nVision:"
                )

                print(
                    image.get(
                        "image_description",
                        ""
                    )[:2000]
                )

        print(
            "\n========== VISUALS =========="
        )

        print(
            f"Visual Pages: "
            f"{len(result['visuals'])}"
        )

        print(
            "\n========== MULTIMODAL KNOWLEDGE =========="
        )

        print(
            result[
                "multimodal_knowledge"
            ][:5000]
        )

    except Exception as e:

        print(
            f"Error: {e}"
        )